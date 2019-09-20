#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import os
from openpyxl import load_workbook
import openpyxl
import arcpy
import pandas as pd
import json
reload(sys)
sys.setdefaultencoding('Windows-1252')

class CreateJson(object):
    def __init__(self):
        BASE_DIR = r'D:\APPS\SURVEY\MG_DGR'
        pathJson = os.path.join(BASE_DIR, 'json')
        self.pathDomain = os.path.join(pathJson, 'domain')
        self.path = os.path.join(BASE_DIR, 'json', 'domain')
        self.wb = load_workbook(os.path.join(BASE_DIR, 'DGR_NotebookMG.xlsx'))
        self.gdb = os.path.join(BASE_DIR, 'SurveyTest.gdb')

    def readSurveyExcelChoices(self):
        ws = self.wb.get_sheet_by_name('choices')
        L = list(set([x[0].value for x in ws['A2':'A10000']]))
        self.listDom = [x for x in L if x is not None]
        l = list(ws.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=3))
        lista = [[x.value for x in i] for i in l]
        self.listaValues = list(filter(lambda x: x != [None, None, None], lista))

    def createJson(self):
        for m in self.listDom:
            pathfile = os.path.join(self.pathDomain, "".join([m, ".json"]))
            file = open(pathfile, "w+")
            text2 = []
            for x in self.listaValues:
                if x[0] == m:
                    if type(x[1])==int:
                        text2.append(u'\n\n{"code":  %s , "name": "%s" }'%(x[1], x[2]))
                    else:
                        text2.append(u'\n\n{"code":  "%s" , "name": "%s" }'%(x[1], x[2]))
            text2 = ",".join(text2)
            text = u'{\n "type": "codedValue", \n "name": "%s", \n "codedValues": [\n\n\n %s ]\n}' % (m, text2)
            file.write(text)
            file.close()
            # print main

    def addJson2Gdb(self):
        listJsons = [os.path.join(self.path, x) for x in os.listdir(self.path)]
        for jsonFile in listJsons:
            with open(jsonFile, "r") as f:
                jx=f.read().decode("Windows-1252")
                df = pd.read_json(jx)
            domain = arcpy.CreateDomain_management(self.gdb, df["name"][0], '', 'TEXT')
            for dom in df['codedValues']:
                val = [v for k,v in dom.items()]
                arcpy.AddCodedValueToDomain_management(self.gdb, df['name'][0], val[0], val[1])

    def main(self):
        self.readSurveyExcelChoices()
        self.createJson()
        # self.addJson2Gdb()

if __name__ == '__main__':
    poo = CreateJson()
    poo.main()
    