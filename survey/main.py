#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import arcpy
import os
import uuid
import pandas as pd
import json
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('Windows-1252')

arcpy.env.overwriteOutput = True
# BASE_DIR = arcpy.GetParameterAsText(0)
# MODULE_NAME = arcpy.GetParameterAsText(1)


BASE_DIR = r'D:\APPS\SURVEY\BDG_DRME_ROCASMENAS'
MODULE_NAME = 'BDG_DRME_ROCASMENAS'
JSONS = os.path.join(BASE_DIR, 'json')
jsonFiles = os.listdir(JSONS)
jsonFiles.remove("domain")
pathDomain = os.path.join(JSONS, "domain")
listDomain = [os.path.join(pathDomain, x) for x in os.listdir(pathDomain)]

class CreateGDB(object):
    def __init__(self):
        self.listFieldsDomain = ["ELEMPRIN", "04_DC_Elemento"], ["ELEMACOM", "04_DC_Elemento"]
        self.wb = load_workbook(os.path.join(BASE_DIR, 'BDG_DRME_ROCASMENAS.xlsx'))
        self.scratch = arcpy.env.scratchGDB

    def createGdb(self):
        """
        Crea la Geodatabase
        :return: 
        """
        self.pathGdb = os.path.join(BASE_DIR, MODULE_NAME + ".gdb")
        self.gdb = arcpy.CreateFileGDB_management(BASE_DIR, MODULE_NAME, "10.0")


    def extractFieldsFromSurvey(self):
        """
        Extrae el tipo de dato, el nombre de la variable y el label de la hoja Survey
        """
        ws = self.wb['survey']
        l = list(ws.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=3))
        lista = [[x.value for x in i] for i in l]
        listaValues = list(filter(lambda x: x != [None, None, None], lista))
        self.listaValuesFeature = []
        for x in listaValues:
            if x[2]!= None:
                m = [x[0], x[1], (x[2].replace("<b>", "")).replace("</b>", "").replace("<i>", "").replace("</i>", "").replace('<u>', "").replace('</u>', "").replace('<font color="blue">', "").replace('</font>', "")]
            else:
                m = x
            self.listaValuesFeature.append(m)

    def extractFieldsFromChoices(self):
        """
        Extrae el nombre del dominio, los valores y la descripcion de la hoja Choices
        """
        ws = self.wb['choices']
        L = list(set([x[0].value for x in ws['A2':'A10000']]))
        self.listDomain = [x for x in L if x is not None]
        l = list(ws.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=3))
        lista = [[x.value for x in i] for i in l]
        self.listaValuesChoices= list(filter(lambda x: x != [None, None, None], lista))
        print self.listaValuesChoices

    def createFeatureJSON(self, ruta, lista):
        prev_a = ""
        count = 1
        diccio = {}
        wsSettings = self.wb['settings']
        firstElm = [["begin repeat", wsSettings["B2"].value, wsSettings["A2"].value]]
        lista2 = firstElm + lista + [["end repeat", None, None]]

        template = text2.append(u'\n\n{"code":  %s , "name": "%s" }' % (x[1], x[2]))


        for i in lista2:
            print i
            if i[0] == "begin repeat" and prev_a == "":
                js = os.path.join(ruta, i[1] + ".json")
                f = open(js, "wb")
                act = f
                diccio[1] = f
            if i[0] == "begin repeat" and prev_a == "begin repeat":
                jg = os.path.join(ruta, i[1] + ".json")
                g = open(jg, "wb")
                act = g
                diccio[2] = g
            if i[0] == "begin repeat" and prev_a == "begin repeatbegin repeat":
                jh = os.path.join(ruta, i[1] + ".json")
                h = open(jh, "wb")
                act = h
                diccio[3] = h
            if i[0] == "begin repeat" and prev_a == "begin repeatbegin repeatbegin repeat":
                jk = os.path.join(ruta, i[1] + ".json")
                k = open(jk, "wb")
                act = k
                diccio[4] = k
            if i[0] == "begin repeat":
                prev_a += i[0]
            elif i[0] == "end repeat":
                prev_a = prev_a[:-12]
            else:
                prev_a = prev_a
            if i[0] == "end repeat":
                act.close()
                act = diccio.get(len(prev_a) / 12)
            if i[0] not in ["begin repeat", "end repeat"]:
                act.write('{}\n'.format(str(i)))
            count += 1

    def createDomainJson(self):
        """
        Crea los JSON para cada dominio
        :return: 
        """
        for m in self.listDomain :
            pathfile = os.path.join(pathDomain, "".join([m, ".json"]))
            file = open(pathfile, "w+")
            text2 = []
            for x in self.listaValuesChoices:
                if x[0] == m:
                    if type(x[1]) == int:
                        text2.append(u'\n\n{"code":  %s , "name": "%s" }' % (x[1], x[2]))
                    else:
                        text2.append(u'\n\n{"code":  "%s" , "name": "%s" }' % (x[1], x[2]))
            text2 = ",".join(text2)
            text = u'{\n "type": "codedValue", \n "name": "%s", \n "codedValues": [\n\n\n %s ]\n}' % (m, text2)
            file.write(text)
            file.close()
            print m

    def fieldsDomain(self):
        """
        Devuelve una lista conteniendo a las variables con el dominio al que se relacionan 
        """
        listFieldsDomain = [[x[1], x[0].split(" ")[1]] for x in self.listaValuesChoices if x[0].split(" ")[0][:6] == "select"]
        print listFieldsDomain

    def addJson2Gdb(self):
        """
        Agregar Dominios a Geodatabase
        :return: 
        """
        listJsons = [os.path.join(pathDomain, x) for x in os.listdir(pathDomain)]
        print listJsons
        for jsonFile in listJsons:
            with open(jsonFile, "r") as f:
                jx = f.read().decode("Windows-1252")
            df = pd.read_json(jx)
            print df
            arcpy.CreateDomain_management(self.gdb, df["name"][0], '', 'TEXT')
            for dom in df['codedValues']:
                val = [v for k, v in dom.items()]
                arcpy.AddCodedValueToDomain_management(self.gdb, df['name'][0], val[0], val[1])


    # def createFeatureJson(self):


    def createFeatures(self):
        for j in jsonFiles:
            name = j.split(".")[-2]
            print name
            jtmp = arcpy.JSONToFeatures_conversion(os.path.join(JSONS, j), os.path.join(self.scratch, name))
            if j.split('_')[0] in ["GPT", "GPL", "GPO"]:
                ft = arcpy.CopyFeatures_management(os.path.join(self.scratch, name), os.path.join(self.pathGdb, name))
            elif j.split('_')[0] == "TB":
                tb = arcpy.JSONToFeatures_conversion(os.path.join(JSONS, j),
                                                     os.path.join(self.scratch, 'a{:.5}'.format(str(uuid.uuid4()))))
                arcpy.TableToTable_conversion(tb, self.gdb, name)


    def assignDomain(self):
        arcpy.env.workspace = self.pathGdb
        for j in jsonFiles:
            name = j.split(".")[-2]
            field_names = [f.name for f in arcpy.ListFields(name)]
            for x in self.listFieldsDomain:
                if x[0] in field_names:
                    arcpy.AssignDomainToField_management(name, x[0], x[1])
            print name

    def process(self):
        # self.createGdb()
        self.extractFieldsFromSurvey()
        # self.extractFieldsFromChoices()
        self.createFeatureJSON(JSONS, self.listaValuesFeature)
        # self.createDomainJson()
        # self.fieldsDomain()
        # self.addJson2Gdb()


    def main(self):
        self.process()

        # print "#--------------------------"
        # print "Domains"
        # self.createDomain()
        # print "#--------------------------"
        # print "Features"
        # self.createFeatures()
        # print "#--------------------------"
        # print "Asignar Domains a features"
        # self.assignDomain()


if __name__ == "__main__":
    poo = CreateGDB()
    poo.main()


# gdb = # Ruta de la gdb
# arcpy.env.workspace = gdb
# lista = arcpy.ListTables("*")
# listaJson = ["".join([x, ".json"]) for x in listaJson]
# for x in lista:
#     arcpy.AddField_management(x, "X","DOUBLE")
#     arcpy.AddField_management(x, "Y","DOUBLE")
#     ft = arcpy.MakeXYEventLayer_management(x, "X", "Y", "in_memory\test")
#     arcpy.FeaturesToJSON_conversion(ft, listaJson[lista.index(x)])

