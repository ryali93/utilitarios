#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import os
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
reload(sys)
sys.setdefaultencoding('Windows-1252')
startTime = datetime.now()
print startTime

BASE_DIR = os.path.join(r"D:\APPS\SURVEY\BDG_DRME_YACMIN")
pathJson = os.path.join(BASE_DIR, "json")

wb = load_workbook(os.path.join(BASE_DIR, "BDG_DRME_YACMIN.xlsx"))
ws=wb.get_sheet_by_name('survey')


L = list(set([x[0].value for x in ws['B2':'B1000']]))
listDom = [x for x in L if x is not None]

l = list(ws.iter_rows(min_row=2, max_row=10000, min_col=1, max_col=3))
lista = [[x.value for x in i] for i in l]

listaValues = list(filter(lambda x: x != [None, None, None], lista))

contador = 1
listaCampos = []
for x in listaValues:
	if x[1].split("_")[0]=="TB":
		with open("{}.json".format(x[1]), "w") as f:




	if (x[1]!=None) and (x[1].split("_")[0]!="GP" or x[1].split("_")[0]!="TB"):
		listaCampos.append(x[1])
print listaCampos

wb.close()


# for m in listDom:
# 	pathfile = os.path.join(path, "".join([m, ".json"]))
# 	file = open(pathfile,"w+")
# 	text2 = []
# 	for x in listaValues:
# 		if x[0] == m:
# 			if type(x[1])==int:
# 				text2.append(u'\n\n{"code":  %s , "name": "%s" }'%(x[1], x[2]))
# 			else:
# 				text2.append(u'\n\n{"code":  "%s" , "name": "%s" }'%(x[1], x[2]))
# 	text2 = ",".join(text2)
# 	text = u'{\n "type": "codedValue", \n "name": "%s", \n "codedValues": [\n\n\n %s ]\n}' % (m, text2)
# 	file.write(text)
# 	file.close()
# 	print m



def main():
    wb = load_workbook(pathExcel)
    ws = wb.get_sheet_by_name("Hoja1")
    createJsons(ws)
    wb.save(pathExcelout)

if __name__ == '__main__':
    main()
    
print datetime.now() - startTime
print "FINALIZADO"


